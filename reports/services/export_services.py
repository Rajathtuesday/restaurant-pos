import csv
import io
import logging
from datetime import timedelta
from decimal import Decimal
from django.db.models import Sum, F
from django.utils import timezone
from orders.models import Order, OrderItem
from core.utils import get_business_date_range
import openpyxl
from openpyxl.styles import Font, Alignment

logger = logging.getLogger("pos.reports")

# Official GST state/UT codes — the first 2 digits of any GSTIN.
# Reference: https://www.gstn.org.in (public, fixed government table).
GST_STATE_CODES = {
    "01": "Jammu and Kashmir", "02": "Himachal Pradesh", "03": "Punjab",
    "04": "Chandigarh", "05": "Uttarakhand", "06": "Haryana", "07": "Delhi",
    "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim",
    "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
    "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
    "24": "Gujarat", "25": "Daman and Diu", "26": "Dadra and Nagar Haveli",
    "27": "Maharashtra", "28": "Andhra Pradesh (Old)", "29": "Karnataka",
    "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
    "34": "Puducherry", "35": "Andaman and Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory",
}


def _place_of_supply(outlet):
    """Real state name derived from the outlet's own GSTIN prefix, instead of
    a hardcoded placeholder. Falls back to a labeled placeholder (not a value
    that reads as real) if the outlet has no GSTIN on file."""
    if outlet and outlet.gst_no and len(outlet.gst_no) >= 2:
        code = outlet.gst_no[:2]
        if code in GST_STATE_CODES:
            return GST_STATE_CODES[code]
    return "Unknown — set outlet GSTIN"

def generate_orders_csv(tenant, outlet, start_date, end_date):
    """Generates a detailed CSV of all orders in the given date range."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Composition Scheme note for CA — only shown when applicable
    if outlet and getattr(outlet, 'is_composition_scheme', False):
        writer.writerow([
            'NOTE: Composition Taxable Person — not eligible to collect tax on supplies.',
            f'GSTIN: {outlet.gst_no or "N/A"}',
            f'Period: {start_date} to {end_date}'
        ])
        writer.writerow([])  # blank separator row

    writer.writerow([
        'Order ID', 'Order No', 'Date', 'Time', 'Outlet', 'Source',
        'Status', 'Customer Name', 'Subtotal', 'Discount', 'GST',
        'Round Off', 'Grand Total', 'Payment Methods'
    ])
    
    range_start, range_end = get_business_date_range(start_date, outlet)
    _, range_end = get_business_date_range(end_date, outlet)
    orders = Order.objects.filter(
        tenant=tenant,
        created_at__gte=range_start,
        created_at__lt=range_end
    ).prefetch_related('payments', 'outlet').order_by('-created_at')

    if outlet:
        orders = orders.filter(outlet=outlet)

    for order in orders:
        payments = ", ".join([p.method for p in order.payments.all()])
        writer.writerow([
            order.id,
            order.order_number or '-',
            order.created_at.strftime('%Y-%m-%d'),
            order.created_at.strftime('%H:%M:%S'),
            order.outlet.name if order.outlet else 'Unknown',
            order.get_source_display(),
            order.get_status_display(),
            order.customer_name or 'Walk-in',
            order.subtotal,
            order.discount_total,
            order.gst_total,
            order.round_off,
            order.grand_total,
            payments
        ])
        
    logger.info("Orders CSV generated successfully. Rows: %s", orders.count())
    return output.getvalue()


def generate_items_csv(tenant, outlet, start_date, end_date):
    """Generates a detailed CSV of all items sold in the given date range."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Item Name', 'Category', 'Quantity Sold', 'Gross Revenue', 'Average Rate'
    ])
    
    # Match the canonical "sold item" definition used by the dashboard
    # (Order.recalculate_totals / item_reports.top_items): only paid/closed
    # orders, exclude voided and complimentary items. The old filter counted
    # items from any order status and included complimentary items, so this
    # export never matched the dashboard's numbers for the same period.
    range_start, _ = get_business_date_range(start_date, outlet)
    _, range_end = get_business_date_range(end_date, outlet)
    items = OrderItem.objects.filter(
        order__tenant=tenant,
        order__created_at__gte=range_start,
        order__created_at__lt=range_end,
        order__status__in=['paid', 'closed'],
        is_complimentary=False,
    ).exclude(status="voided")

    if outlet:
        items = items.filter(order__outlet=outlet)

    item_stats = items.values(
        'menu_item__name', 'menu_item__category__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_rev=Sum('total_price')
    ).order_by('-total_qty')
    
    for stat in item_stats:
        qty = stat['total_qty'] or 0
        rev = stat['total_rev'] or 0
        avg_rate = (rev / qty) if qty > 0 else 0
        writer.writerow([
            stat['menu_item__name'] or 'Unknown Item',
            stat['menu_item__category__name'] or 'Uncategorized',
            qty,
            rev,
            round(avg_rate, 2)
        ])
        
    logger.info("Items CSV generated successfully. Unique Items: %s", item_stats.count())
    return output.getvalue()


def _autosize_columns(ws):
    """Auto-adjust column widths (skip MergedCells which lack column_letter)."""
    for col in ws.columns:
        max_length = 0
        first_cell = col[0]
        if not hasattr(first_cell, "column_letter"):
            continue
        column = first_cell.column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2


def generate_gstr1_excel(tenant, outlet, start_date, end_date):
    """
    Generates a GSTR-1 compliant Excel report for B2C sales, plus the
    Table 12 HSN/SAC summary — mandatory for every GST filer regardless
    of B2B/B2C mix, unlike the B2CS sheet which only matters once there's
    B2C turnover to report.
    """
    from orders.services.tax_service import split_cgst_sgst

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-1 B2CS"
    
    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"GSTR-1 B2CS (Business to Consumer Small) Report - {tenant.name}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    date_cell = ws['A2']
    date_cell.value = f"Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = [
        'Type', 'Place of Supply', 'Rate (%)', 'Taxable Value', 
        'Central Tax (CGST)', 'State Tax (SGST)', 'Integrated Tax (IGST)', 
        'Cess Amount', 'E-Commerce GSTIN', 'Total Tax'
    ]
    
    ws.append([]) # Empty row
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True)
        
    # We group by the menu item's GST percentage
    # To handle order-level discounts, we calculate an approximation:
    # Taxable Value = Sum(total_price)
    # GST = Sum(GST calculated on item base considering order discounts)
    # For a perfect GSTR-1, we will re-calculate item-level taxable amounts accounting for order discounts
    
    # Let's fetch the actual orders to iterate through their items to be perfectly accurate with order discounts
    range_start, _ = get_business_date_range(start_date, outlet)
    _, range_end = get_business_date_range(end_date, outlet)
    orders = Order.objects.filter(
        tenant=tenant,
        created_at__gte=range_start,
        created_at__lt=range_end,
        status__in=['paid', 'closed']
    ).prefetch_related('items', 'items__menu_item')
    
    if outlet:
        orders = orders.filter(outlet=outlet)
        
    # Group data by GST Rate
    gst_groups = {} # rate: {'taxable': 0, 'gst': 0}
    
    for order in orders:
        # Composition scheme outlets issue Bill of Supply — no GST, skip from GSTR-1
        if getattr(order.outlet, "is_composition_scheme", False):
            continue

        gst_inclusive = getattr(order.outlet, "gst_inclusive", False)

        # Reconstruct exactly how Order.recalculate_totals works
        items_valid = [item for item in order.items.all() if item.status != "voided" and not item.is_complimentary]
        
        raw_subtotal = sum((item.total_price for item in items_valid), Decimal("0.0"))
        
        item_discount_total = Decimal("0.00")
        subtotal_after_item_discounts = Decimal("0.00")
        for item in items_valid:
            item_base = item.total_price
            if getattr(item, 'item_discount_pct', Decimal("0.00")) > 0:
                item_discount = item_base * (item.item_discount_pct / Decimal("100"))
                item_base = item_base - item_discount
            subtotal_after_item_discounts += item_base
            
        order_discount_total = Decimal("0.00")
        if order.discount_type == "percentage" and (order.discount_value or 0) > 0:
            order_discount_total = subtotal_after_item_discounts * (Decimal(order.discount_value) / Decimal("100"))
        elif order.discount_type == "amount" and (order.discount_value or 0) > 0:
            order_discount_total = Decimal(str(order.discount_value))
            
        if subtotal_after_item_discounts > 0:
            order_discount_factor = max(Decimal("0.0"), (subtotal_after_item_discounts - order_discount_total) / subtotal_after_item_discounts)
        else:
            order_discount_factor = Decimal("1.0")
            
        for item in items_valid:
            item_base = item.total_price
            if getattr(item, 'item_discount_pct', Decimal("0.00")) > 0:
                item_base = item_base * (1 - item.item_discount_pct / Decimal("100"))
            
            item_discounted = item_base * order_discount_factor
            rate = item.menu_item.gst_percentage if item.menu_item else Decimal("5.00")

            rate_key = float(rate)
            if rate_key not in gst_groups:
                gst_groups[rate_key] = {'taxable': Decimal("0.0"), 'gst': Decimal("0.0")}

            if gst_inclusive:
                # item_discounted is the inclusive customer price — back-calculate
                item_gst     = item_discounted * rate / (Decimal("100") + rate)
                item_taxable = item_discounted - item_gst
            else:
                # item_discounted is already the pre-GST base
                item_taxable = item_discounted
                item_gst     = item_taxable * rate / Decimal("100")

            gst_groups[rate_key]['taxable'] += item_taxable
            gst_groups[rate_key]['gst']     += item_gst

    total_taxable = Decimal("0.0")
    total_central = Decimal("0.0")
    total_state = Decimal("0.0")

    # Real place of supply, derived from the outlet's own GSTIN — not a
    # hardcoded placeholder. If the report spans multiple outlets that don't
    # share one state, don't guess: this report only computes CGST/SGST
    # (intra-state), never IGST, so mixing states into one "Local" row would
    # be silently wrong for whichever outlet isn't actually local. Proper
    # multi-state GSTR-1 (with real IGST math) is a larger, separate piece of
    # work — this only fixes the misleading placeholder value.
    if outlet:
        pos_state = _place_of_supply(outlet)
    else:
        outlet_states = {o.gst_no[:2] for o in tenant.outlets.all() if o.gst_no and len(o.gst_no) >= 2}
        if len(outlet_states) == 1:
            pos_state = _place_of_supply(tenant.outlets.filter(gst_no__startswith=next(iter(outlet_states))).first())
        elif len(outlet_states) > 1:
            pos_state = "Multiple states — export per outlet"
            logger.warning(
                "GSTR-1 export for tenant %s spans outlets in multiple states (%s); "
                "this report only computes CGST/SGST, not IGST. Export per outlet instead.",
                tenant.id, outlet_states,
            )
        else:
            pos_state = "Unknown — set outlet GSTIN"
    
    for rate, data in sorted(gst_groups.items()):
        taxable = data['taxable']
        gst = data['gst']

        cgst, sgst = split_cgst_sgst(gst)

        total_taxable += taxable
        total_central += cgst
        total_state += sgst
        
        ws.append([
            'OE', # Outward Supplies
            pos_state,
            rate,
            round(float(taxable), 2),
            round(float(cgst), 2),
            round(float(sgst), 2),
            0.0, # IGST
            0.0, # Cess
            '',  # E-Comm GSTIN
            round(float(gst), 2)
        ])
        
    # Totals Row
    ws.append([])
    ws.append([
        'TOTAL', '', '', 
        round(float(total_taxable), 2),
        round(float(total_central), 2),
        round(float(total_state), 2),
        0.0, 0.0, '', 
        round(float(total_central + total_state), 2)
    ])
    
    # Bold the totals row
    for col in range(1, 11):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    _autosize_columns(ws)

    # ── Table 12 — HSN/SAC Summary ──────────────────────────────────────────
    # Mandatory for every GSTR-1 filer, unconditionally — unlike the B2CS
    # sheet above, this isn't optional just because a period had no B2C sales.
    # A restaurant's whole menu is one GST service classification (SAC 996331,
    # "restaurant/catering services"), so this reuses gst_groups computed
    # above rather than re-deriving anything — one row per rate actually used.
    ws12 = wb.create_sheet("GSTR-1 Table 12 (HSN)")

    ws12.merge_cells('A1:K1')
    t12_title = ws12['A1']
    t12_title.value = f"GSTR-1 Table 12 — HSN/SAC Summary - {tenant.name}"
    t12_title.font = Font(size=14, bold=True)
    t12_title.alignment = Alignment(horizontal='center')

    ws12.merge_cells('A2:K2')
    t12_date = ws12['A2']
    t12_date.value = f"Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    t12_date.font = Font(italic=True)
    t12_date.alignment = Alignment(horizontal='center')

    t12_headers = [
        'HSN/SAC', 'Description', 'UQC', 'Total Quantity', 'Total Value',
        'Rate (%)', 'Taxable Value', 'Integrated Tax (IGST)',
        'Central Tax (CGST)', 'State Tax (SGST)', 'Cess Amount',
    ]
    ws12.append([])
    ws12.append(t12_headers)
    for col in range(1, len(t12_headers) + 1):
        ws12.cell(row=4, column=col).font = Font(bold=True)

    RESTAURANT_SAC = "996331"
    RESTAURANT_SAC_DESC = "Restaurant/catering services"

    t12_total_value = Decimal("0.0")
    t12_total_taxable = Decimal("0.0")
    t12_total_central = Decimal("0.0")
    t12_total_state = Decimal("0.0")

    for rate, data in sorted(gst_groups.items()):
        taxable = data['taxable']
        gst = data['gst']
        cgst, sgst = split_cgst_sgst(gst)
        total_value = taxable + gst

        t12_total_value += total_value
        t12_total_taxable += taxable
        t12_total_central += cgst
        t12_total_state += sgst

        ws12.append([
            RESTAURANT_SAC,
            RESTAURANT_SAC_DESC,
            'NA',   # services have no unit of measure
            0,      # Total Quantity — not applicable for services
            round(float(total_value), 2),
            rate,
            round(float(taxable), 2),
            0.0,    # IGST — same intra-state-only limitation as the B2CS sheet
            round(float(cgst), 2),
            round(float(sgst), 2),
            0.0,    # Cess
        ])

    ws12.append([])
    ws12.append([
        'TOTAL', '', '', 0,
        round(float(t12_total_value), 2), '',
        round(float(t12_total_taxable), 2),
        0.0,
        round(float(t12_total_central), 2),
        round(float(t12_total_state), 2),
        0.0,
    ])
    for col in range(1, len(t12_headers) + 1):
        ws12.cell(row=ws12.max_row, column=col).font = Font(bold=True)

    _autosize_columns(ws12)

    output = io.BytesIO()
    wb.save(output)
    
    logger.info("GSTR-1 Excel generated successfully. Max Row: %s", ws.max_row)
    return output.getvalue()


def generate_waiter_csv(tenant, outlet, start_date, end_date):
    """Generates Staff Performance CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Staff Name', 'Total Orders Handled', 'Total Revenue Handled', 'Average Order Value'])
    
    range_start, _ = get_business_date_range(start_date, outlet)
    _, range_end = get_business_date_range(end_date, outlet)
    orders = Order.objects.filter(
        tenant=tenant,
        created_at__gte=range_start,
        created_at__lt=range_end,
        status__in=['paid', 'closed'],
        created_by__isnull=False
    )
    
    if outlet:
        orders = orders.filter(outlet=outlet)
        
    waiter_stats = orders.values('created_by__username').annotate(
        total_orders=Sum(1),
        total_rev=Sum('grand_total')
    ).order_by('-total_rev')
    
    for stat in waiter_stats:
        orders_count = stat['total_orders'] or 0
        rev = stat['total_rev'] or 0
        aov = (rev / orders_count) if orders_count > 0 else 0
        writer.writerow([
            stat['created_by__username'],
            orders_count,
            round(float(rev), 2),
            round(float(aov), 2)
        ])
        
    logger.info("Waiter Performance CSV generated successfully. Waiters analyzed: %s", waiter_stats.count())
    return output.getvalue()


def generate_category_csv(tenant, outlet, start_date, end_date):
    """Generates Category Sales CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Category Name', 'Items Sold', 'Total Revenue'])
    
    # Same canonical "sold item" definition as the dashboard — exclude voided
    # and complimentary items so this export agrees with the on-screen numbers.
    range_start, _ = get_business_date_range(start_date, outlet)
    _, range_end = get_business_date_range(end_date, outlet)
    items = OrderItem.objects.filter(
        order__tenant=tenant,
        order__created_at__gte=range_start,
        order__created_at__lt=range_end,
        order__status__in=['paid', 'closed'],
        is_complimentary=False,
    ).exclude(status="voided")

    if outlet:
        items = items.filter(order__outlet=outlet)

    category_stats = items.values('menu_item__category__name').annotate(
        qty=Sum('quantity'),
        rev=Sum('total_price')
    ).order_by('-rev')
    
    for stat in category_stats:
        writer.writerow([
            stat['menu_item__category__name'] or 'Uncategorized',
            stat['qty'] or 0,
            round(float(stat['rev'] or 0), 2)
        ])
        
    logger.info("Category Sales CSV generated successfully. Categories analyzed: %s", category_stats.count())
    return output.getvalue()


def generate_pl_csv(tenant, outlet, start_date, end_date):
    """Generates the Net Profit / P&L CSV — revenue, COGS, operating
    expenses, net profit, and the expense category breakdown."""
    from reports.services.pl_reports import net_profit_report

    output = io.StringIO()
    writer = csv.writer(output)
    report = net_profit_report(tenant, outlet, start_date, end_date)

    writer.writerow(['NET PROFIT REPORT', f'Period: {start_date} to {end_date}'])
    writer.writerow([])
    writer.writerow(['Gross Revenue', report['gross_revenue']])
    writer.writerow(['GST Collected', report['gst_collected']])
    writer.writerow(['Net Revenue', report['net_revenue']])
    writer.writerow(['Discounts Given', report['discounts']])
    writer.writerow(['COGS', report['cogs']])
    writer.writerow(['Gross Profit', report['gross_profit']])
    writer.writerow(['Gross Margin %', report['gross_margin_pct']])
    writer.writerow(['Operating Expenses', report['operating_expenses']])
    writer.writerow(['NET PROFIT', report['net_profit']])
    writer.writerow(['Net Margin %', report['net_margin_pct']])
    writer.writerow([])
    writer.writerow(['EXPENSE BREAKDOWN BY CATEGORY'])
    writer.writerow(['Category', 'Amount'])
    for row in report['expense_breakdown']:
        writer.writerow([row['category'], float(row['total'])])

    return output.getvalue()


def generate_menu_engineering_csv(tenant, outlet, start_date, end_date):
    """Generates the Menu Engineering (stars/dogs quadrant) CSV."""
    from reports.services.menu_engineering import menu_engineering_report

    output = io.StringIO()
    writer = csv.writer(output)
    rows = menu_engineering_report(tenant, outlet, start_date, end_date)

    writer.writerow(['Item', 'Quantity Sold', 'Revenue', 'COGS', 'Margin %', 'Quadrant', 'Cost Known'])
    for row in rows["items"]:
        writer.writerow([
            row['name'], row['qty'], row['revenue'],
            row['cogs'] if row['cogs_known'] else 'unknown',
            row['margin_pct'] if row['cogs_known'] else '',
            row['quadrant'], 'yes' if row['cogs_known'] else 'no',
        ])

    return output.getvalue()


def generate_labor_csv(tenant, outlet, start_date, end_date):
    """Generates the Labor Cost CSV -- per-staff hours, tips, and cost."""
    from reports.services.labor_reports import labor_cost_report

    output = io.StringIO()
    writer = csv.writer(output)
    report = labor_cost_report(tenant, outlet, start_date, end_date)

    writer.writerow(['LABOR COST REPORT', f'Period: {start_date} to {end_date}'])
    writer.writerow(['Total Labor Cost', report['total_labor_cost']])
    writer.writerow(['Revenue', report.get('revenue', 0)])
    writer.writerow(['Labor Cost %', report['labor_cost_pct']])
    writer.writerow([])
    writer.writerow(['Staff', 'Pay Type', 'Hours', 'Tips', 'Cost'])
    for row in report['rows']:
        writer.writerow([
            row['username'], row['pay_type'] or 'unknown', row['hours'], row['tips'],
            row['cost'] if row['cost_known'] else 'unknown',
        ])

    return output.getvalue()


def generate_audit_csv(tenant, outlet, start_date, end_date):
    """Generates the Discount/Void Staff Audit CSV."""
    from reports.services.audit_reports import discount_void_audit

    output = io.StringIO()
    writer = csv.writer(output)
    report = discount_void_audit(tenant, outlet, start_date, end_date)

    writer.writerow(['DISCOUNT / VOID STAFF AUDIT', f'Period: {start_date} to {end_date}'])
    writer.writerow([])

    writer.writerow(['ORDER-LEVEL DISCOUNTS'])
    writer.writerow(['Staff', 'Count'])
    for row in report['discounts']:
        writer.writerow([row['created_by__username'] or 'Unknown', row['count']])
    writer.writerow([])

    writer.writerow(['ITEM-LEVEL DISCOUNTS (from rollout date onward)'])
    writer.writerow(['Staff', 'Count'])
    for row in report['item_discounts']:
        writer.writerow([row['created_by__username'] or 'Unknown', row['count']])
    writer.writerow([])

    writer.writerow(['COMPLIMENTARY ITEMS (from rollout date onward)'])
    writer.writerow(['Staff', 'Count'])
    for row in report['comps']:
        writer.writerow([row['created_by__username'] or 'Unknown', row['count']])
    writer.writerow([])

    writer.writerow(['ITEM VOIDS'])
    writer.writerow(['Staff', 'Count'])
    for row in report['voids']:
        writer.writerow([row['created_by__username'] or 'Unknown', row['count']])
    writer.writerow([])

    writer.writerow(['VOID REASONS'])
    writer.writerow(['Reason', 'Count'])
    for row in report['void_reasons']:
        writer.writerow([row['metadata__reason'] or 'Unspecified', row['count']])

    return output.getvalue()


def generate_crm_analytics_csv(tenant, outlet, start_date, end_date):
    """Generates the CRM/Loyalty Analytics CSV."""
    from reports.services.crm_reports import crm_analytics_report

    output = io.StringIO()
    writer = csv.writer(output)
    report = crm_analytics_report(tenant, outlet, start_date, end_date)

    writer.writerow(['CRM / LOYALTY ANALYTICS', f'Period: {start_date} to {end_date}'])
    writer.writerow(['Repeat Customer Rate %', report['repeat_rate_pct']])
    writer.writerow(['Repeat Guests', report['repeat_guests']])
    writer.writerow(['Active Guests', report['active_guests']])
    writer.writerow(['Average Rating', report['avg_rating'] if report['avg_rating'] is not None else 'n/a'])
    writer.writerow([])

    writer.writerow(['LOYALTY TREND'])
    writer.writerow(['Date', 'Type', 'Points', 'Count'])
    for row in report['loyalty_trend']:
        writer.writerow([row['day'], row['transaction_type'], row['points'], row['count']])
    writer.writerow([])

    writer.writerow(['FEEDBACK TREND'])
    writer.writerow(['Date', 'Avg Rating', 'Count'])
    for row in report['feedback_trend']:
        writer.writerow([row['day'], round(row['avg_rating'], 2), row['count']])

    return output.getvalue()
