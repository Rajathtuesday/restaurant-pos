# reports/tests/test_reports.py
"""
Test suite for the reports application.
Covers:
  - reports.services.sales_reports: daily_sales, hourly_sales
  - reports.services.item_reports: top_items
  - reports.services.kitchen_reports: kitchen_performance, top_kitchen_items
  - reports.services.export_services: CSV and Excel generation
  - reports.views: dashboard and export_reports HTTP endpoints
"""

import io
import csv
from decimal import Decimal
from django.test import TestCase, Client
from django.urls import reverse
from django.utils import timezone

from tenants.models import Tenant, Outlet
from accounts.models import User
from menu.models import MenuCategory, MenuItem
from orders.models import Order, OrderItem, Payment
from core.utils import get_business_date


# ---------------------------------------------------------------------------
# SHARED FIXTURE HELPER
# ---------------------------------------------------------------------------

def create_base_fixtures():
    """Creates reusable Tenant, Outlet, User, and MenuItem objects."""
    tenant = Tenant.objects.create(name="Test Restaurant")
    outlet = Outlet.objects.create(tenant=tenant, name="Main Branch")
    user = User.objects.create_user(
        username="owner_test",
        password="pass1234",
        tenant=tenant,
        outlet=outlet,
        role="owner"
    )
    category = MenuCategory.objects.create(
        tenant=tenant,
        outlet=outlet,
        name="Starters"
    )
    item = MenuItem.objects.create(
        tenant=tenant,
        outlet=outlet,
        category=category,
        name="Paneer Tikka",
        price=Decimal("200.00"),
        gst_percentage=Decimal("5.00"),
        is_available=True
    )
    return tenant, outlet, user, item


def create_paid_order(tenant, outlet, user, item, grand_total=500, method="cash"):
    """Creates a paid order with an associated payment."""
    order = Order.objects.create(
        tenant=tenant,
        outlet=outlet,
        created_by=user,
        status="paid",
        subtotal=Decimal(str(grand_total)),
        gst_total=Decimal("0.00"),
        discount_total=Decimal("0.00"),
        grand_total=Decimal(str(grand_total))
    )
    Payment.objects.create(
        order=order,
        method=method,
        amount=Decimal(str(grand_total)),
        created_by=user
    )
    if item:
        OrderItem.objects.create(
            order=order,
            menu_item=item,
            quantity=2,
            price=item.price,
            gst_percentage=item.gst_percentage,
            total_price=Decimal("400.00"),
            status="served"
        )
    return order


# ---------------------------------------------------------------------------
# 1. SALES REPORT SERVICE TESTS
# ---------------------------------------------------------------------------

class DailySalesServiceTest(TestCase):
    """Tests for reports.services.sales_reports.daily_sales"""

    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()

    def test_daily_sales_returns_correct_total(self):
        """Revenue should sum all payments for today."""
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=500)
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=300)

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        self.assertEqual(result["total_sales"], 800.0)

    def test_daily_sales_zero_when_no_orders(self):
        """Revenue should be 0 when no orders exist."""
        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        self.assertEqual(result["total_sales"], 0.0)
        self.assertEqual(result["orders"], 0)

    def test_daily_sales_counts_orders_correctly(self):
        """Order count should match the number of paid orders."""
        create_paid_order(self.tenant, self.outlet, self.user, self.item)
        create_paid_order(self.tenant, self.outlet, self.user, self.item)

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        self.assertEqual(result["orders"], 2)

    def test_daily_sales_isolates_by_outlet(self):
        """Sales for outlet A must not bleed into outlet B."""
        other_outlet = Outlet.objects.create(tenant=self.tenant, name="Branch B")
        other_user = User.objects.create_user(
            username="owner_b", password="pass", tenant=self.tenant,
            outlet=other_outlet, role="owner"
        )
        create_paid_order(self.tenant, other_outlet, other_user, self.item, grand_total=999)
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=200)

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        self.assertEqual(result["total_sales"], 200.0)

    def test_daily_sales_average_order_value(self):
        """AOV should be total_sales / order_count."""
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=300)
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=700)

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        self.assertEqual(result["avg_order_value"], 500.0)

    def test_daily_sales_excludes_refunds_from_payment_split(self):
        """Refund payments should appear as 'net_refunds', not in payment methods."""
        order = create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=500)
        Payment.objects.create(
            order=order, method="refund",
            amount=Decimal("-200.00"), created_by=self.user
        )

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        methods = [p["method"] for p in result["payments"]]
        self.assertNotIn("refund", methods)
        self.assertEqual(result["net_refunds"], 200.0)

    def test_daily_sales_correct_payment_split(self):
        """Payment split should show the correct totals per method."""
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=400, method="cash")
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=600, method="upi")

        from reports.services.sales_reports import daily_sales
        result = daily_sales(self.tenant, self.outlet)

        split = {p["method"]: p["total"] for p in result["payments"]}
        self.assertEqual(float(split["cash"]), 400.0)
        self.assertEqual(float(split["upi"]), 600.0)


# ---------------------------------------------------------------------------
# 2. ITEM REPORT SERVICE TESTS
# ---------------------------------------------------------------------------

class TopItemsServiceTest(TestCase):
    """Tests for reports.services.item_reports.top_items"""

    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()

    def test_top_items_returns_items(self):
        """top_items should return sold items."""
        create_paid_order(self.tenant, self.outlet, self.user, self.item)

        from reports.services.item_reports import top_items
        result = list(top_items(self.tenant, self.outlet))

        self.assertTrue(len(result) > 0)
        self.assertEqual(result[0]["menu_item__name"], "Paneer Tikka")

    def test_top_items_empty_when_no_orders(self):
        """top_items should be empty when no orders exist."""
        from reports.services.item_reports import top_items
        result = list(top_items(self.tenant, self.outlet))

        self.assertEqual(len(result), 0)

    def test_top_items_orders_by_quantity_descending(self):
        """Most sold item should appear first."""
        category = MenuCategory.objects.get(tenant=self.tenant)
        item2 = MenuItem.objects.create(
            tenant=self.tenant, outlet=self.outlet, category=category,
            name="Chicken Wings", price=Decimal("300.00"),
            gst_percentage=Decimal("5.00"), is_available=True
        )
        order1 = create_paid_order(self.tenant, self.outlet, self.user, self.item)
        order2 = create_paid_order(self.tenant, self.outlet, self.user, None, grand_total=600)
        # Add 5 wings to order2 to make it the top item
        for _ in range(3):
            OrderItem.objects.create(
                order=order2, menu_item=item2, quantity=5,
                price=item2.price, gst_percentage=item2.gst_percentage,
                total_price=Decimal("1500.00"), status="served"
            )

        from reports.services.item_reports import top_items
        result = list(top_items(self.tenant, self.outlet))

        self.assertEqual(result[0]["menu_item__name"], "Chicken Wings")


# ---------------------------------------------------------------------------
# 3. KITCHEN REPORT SERVICE TESTS
# ---------------------------------------------------------------------------

class KitchenReportServiceTest(TestCase):
    """Tests for reports.services.kitchen_reports"""

    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()

    def test_kitchen_performance_returns_zeroes_when_empty(self):
        """kitchen_performance should return zeros when no kitchen data exists."""
        from reports.services.kitchen_reports import kitchen_performance
        result = kitchen_performance(self.tenant, self.outlet)

        self.assertEqual(result["total_items_prepared"], 0)
        self.assertEqual(result["total_kots"], 0)

    def test_top_kitchen_items_empty_when_no_kots(self):
        """top_kitchen_items should return empty list when no KOT has been sent."""
        from reports.services.kitchen_reports import top_kitchen_items
        result = top_kitchen_items(self.tenant, self.outlet)

        self.assertEqual(result, [])


# ---------------------------------------------------------------------------
# 4. EXPORT SERVICE TESTS
# ---------------------------------------------------------------------------

class ExportServicesTest(TestCase):
    """Tests for reports.services.export_services"""

    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()
        # The export services bound their query with get_business_date_range(),
        # cutoff-aware (default 6 AM), not calendar-date-aware. Plain
        # timezone.localdate() disagrees with that window for any order
        # created between midnight and the cutoff.
        self.today = get_business_date(timezone.now(), self.outlet)
        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=500)

    def test_generate_orders_csv_returns_string(self):
        """Should return a non-empty CSV string."""
        from reports.services.export_services import generate_orders_csv
        result = generate_orders_csv(self.tenant, self.outlet, self.today, self.today)

        self.assertIsInstance(result, str)
        self.assertGreater(len(result), 0)

    def test_generate_orders_csv_has_correct_headers(self):
        """CSV should contain expected column headers."""
        from reports.services.export_services import generate_orders_csv
        result = generate_orders_csv(self.tenant, self.outlet, self.today, self.today)
        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        self.assertIn("Order ID", headers)
        self.assertIn("Order No", headers)
        self.assertIn("Grand Total", headers)

    def test_generate_orders_csv_contains_order_data(self):
        """CSV data rows should reflect the created order."""
        from reports.services.export_services import generate_orders_csv
        result = generate_orders_csv(self.tenant, self.outlet, self.today, self.today)
        reader = csv.reader(io.StringIO(result))
        next(reader)  # skip header
        rows = list(reader)

        self.assertEqual(len(rows), 1)
        # Grand total column (index 12) should be 500
        self.assertIn("500", rows[0][12])

    def test_generate_items_csv_returns_string(self):
        """Should return a non-empty CSV string."""
        from reports.services.export_services import generate_items_csv
        result = generate_items_csv(self.tenant, self.outlet, self.today, self.today)

        self.assertIsInstance(result, str)
        self.assertGreater(len(result), 0)

    def test_generate_items_csv_has_correct_headers(self):
        """Items CSV should contain expected columns."""
        from reports.services.export_services import generate_items_csv
        result = generate_items_csv(self.tenant, self.outlet, self.today, self.today)
        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        self.assertIn("Item Name", headers)
        self.assertIn("Quantity Sold", headers)
        self.assertIn("Average Rate", headers)

    def test_generate_items_csv_contains_item_data(self):
        """Items CSV should contain the sold item name."""
        from reports.services.export_services import generate_items_csv
        result = generate_items_csv(self.tenant, self.outlet, self.today, self.today)

        self.assertIn("Paneer Tikka", result)

    def test_generate_waiter_csv_returns_string(self):
        """Waiter CSV should return a non-empty string."""
        from reports.services.export_services import generate_waiter_csv
        result = generate_waiter_csv(self.tenant, self.outlet, self.today, self.today)

        self.assertIsInstance(result, str)
        self.assertGreater(len(result), 0)

    def test_generate_waiter_csv_has_correct_headers(self):
        """Waiter CSV should contain expected columns."""
        from reports.services.export_services import generate_waiter_csv
        result = generate_waiter_csv(self.tenant, self.outlet, self.today, self.today)
        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        self.assertIn("Staff Name", headers)
        self.assertIn("Total Revenue Handled", headers)
        self.assertIn("Average Order Value", headers)

    def test_generate_category_csv_returns_string(self):
        """Category CSV should return a non-empty string."""
        from reports.services.export_services import generate_category_csv
        result = generate_category_csv(self.tenant, self.outlet, self.today, self.today)

        self.assertIsInstance(result, str)
        self.assertGreater(len(result), 0)

    def test_generate_category_csv_has_correct_headers(self):
        """Category CSV should contain expected columns."""
        from reports.services.export_services import generate_category_csv
        result = generate_category_csv(self.tenant, self.outlet, self.today, self.today)
        reader = csv.reader(io.StringIO(result))
        headers = next(reader)

        self.assertIn("Category Name", headers)
        self.assertIn("Items Sold", headers)
        self.assertIn("Total Revenue", headers)

    def test_generate_gstr1_excel_returns_bytes(self):
        """GSTR-1 should return a bytes object (valid .xlsx binary)."""
        from reports.services.export_services import generate_gstr1_excel
        result = generate_gstr1_excel(self.tenant, self.outlet, self.today, self.today)

        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_generate_gstr1_excel_is_valid_xlsx(self):
        """GSTR-1 bytes should be loadable as a valid openpyxl workbook with correct structure."""
        import openpyxl
        from reports.services.export_services import generate_gstr1_excel
        result = generate_gstr1_excel(self.tenant, self.outlet, self.today, self.today)
        wb = openpyxl.load_workbook(io.BytesIO(result))

        self.assertIsNotNone(wb)
        ws = wb.active

        # Row 1 is the merged title cell
        self.assertIn(self.tenant.name, ws['A1'].value)

        # Row 4 is the header row (row1=title, row2=period, row3=empty, row4=headers)
        header_values = [ws.cell(row=4, column=i).value for i in range(1, 11)]
        self.assertIn("Rate (%)", header_values)
        self.assertIn("Central Tax (CGST)", header_values)
        self.assertIn("State Tax (SGST)", header_values)
        self.assertIn("Taxable Value", header_values)

    def test_generate_gstr1_excel_includes_hsn_summary_sheet(self):
        """Table 12 (HSN/SAC summary) is mandatory for every GSTR-1 filer,
        not just tenants with B2B sales — it must always be present, with
        the fixture's known 5% item correctly rolled into SAC 996331."""
        import openpyxl
        from reports.services.export_services import generate_gstr1_excel
        result = generate_gstr1_excel(self.tenant, self.outlet, self.today, self.today)
        wb = openpyxl.load_workbook(io.BytesIO(result))

        self.assertIn("GSTR-1 Table 12 (HSN)", wb.sheetnames)
        ws12 = wb["GSTR-1 Table 12 (HSN)"]

        header_values = [ws12.cell(row=4, column=i).value for i in range(1, 12)]
        self.assertIn("HSN/SAC", header_values)
        self.assertIn("UQC", header_values)
        self.assertIn("Taxable Value", header_values)
        self.assertIn("Central Tax (CGST)", header_values)
        self.assertIn("State Tax (SGST)", header_values)

        # Row 5 is the first (and only, one rate group) data row.
        data_row = [ws12.cell(row=5, column=i).value for i in range(1, 12)]
        self.assertEqual(data_row[0], "996331")   # HSN/SAC
        self.assertEqual(data_row[5], 5.0)        # Rate (%)
        self.assertEqual(data_row[6], 400.0)      # Taxable Value
        self.assertEqual(data_row[8], 10.0)       # CGST
        self.assertEqual(data_row[9], 10.0)       # SGST

    def test_generate_orders_csv_empty_when_no_orders_in_range(self):
        """CSV with a past date range should only have the header, no data rows."""
        from reports.services.export_services import generate_orders_csv
        from datetime import date
        result = generate_orders_csv(self.tenant, self.outlet, date(2000, 1, 1), date(2000, 1, 2))
        reader = csv.reader(io.StringIO(result))
        next(reader)  # header
        rows = list(reader)
        self.assertEqual(len(rows), 0)


# ---------------------------------------------------------------------------
# 5. REPORTS VIEW (HTTP ENDPOINT) TESTS
# ---------------------------------------------------------------------------

class ReportsDashboardViewTest(TestCase):
    """Tests for reports.views.dashboard (HTTP)"""

    def setUp(self):
        self.client = Client()
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()

    def test_dashboard_requires_login(self):
        """Unauthenticated users should be redirected to login."""
        response = self.client.get("/reports/dashboard/")
        self.assertIn(response.status_code, [302, 403])

    def test_dashboard_loads_for_owner(self):
        """Authenticated owner should get HTTP 200."""
        self.client.login(username="owner_test", password="pass1234")
        response = self.client.get("/reports/dashboard/")
        self.assertEqual(response.status_code, 200)

    def test_dashboard_forbidden_for_staff_without_role(self):
        """A user with no special role should be forbidden."""
        staff = User.objects.create_user(
            username="plain_staff", password="pass",
            tenant=self.tenant, outlet=self.outlet, role="staff"
        )
        self.client.login(username="plain_staff", password="pass")
        response = self.client.get("/reports/dashboard/")
        self.assertEqual(response.status_code, 403)


class ExportReportsViewTest(TestCase):
    """Tests for reports.views.export_reports (HTTP)"""

    def setUp(self):
        self.client = Client()
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()
        self.today = get_business_date(timezone.now(), self.outlet).isoformat()
        create_paid_order(self.tenant, self.outlet, self.user, self.item)
        self.client.login(username="owner_test", password="pass1234")

    def _get_export(self, export_type):
        return self.client.get(
            f"/reports/export/?type={export_type}&date_filter=today"
        )

    def test_export_orders_csv_returns_200(self):
        """Orders CSV export endpoint should respond with HTTP 200."""
        response = self._get_export("orders")
        self.assertEqual(response.status_code, 200)

    def test_export_orders_csv_content_type(self):
        """Orders export should return text/csv content-type."""
        response = self._get_export("orders")
        self.assertIn("text/csv", response["Content-Type"])

    def test_export_orders_csv_has_attachment_header(self):
        """Orders export should have Content-Disposition attachment header."""
        response = self._get_export("orders")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".csv", response["Content-Disposition"])

    def test_export_items_csv_returns_200(self):
        """Items CSV export should return HTTP 200."""
        response = self._get_export("items")
        self.assertEqual(response.status_code, 200)

    def test_export_waiters_csv_returns_200(self):
        """Waiters CSV export should return HTTP 200."""
        response = self._get_export("waiters")
        self.assertEqual(response.status_code, 200)

    def test_export_categories_csv_returns_200(self):
        """Categories CSV export should return HTTP 200."""
        response = self._get_export("categories")
        self.assertEqual(response.status_code, 200)

    def test_export_gstr1_excel_returns_200(self):
        """GSTR-1 Excel export should return HTTP 200."""
        response = self._get_export("gstr1")
        self.assertEqual(response.status_code, 200)

    def test_export_gstr1_content_type_is_excel(self):
        """GSTR-1 export should return xlsx content-type."""
        response = self._get_export("gstr1")
        self.assertIn(
            "spreadsheetml",
            response["Content-Type"]
        )

    def test_export_invalid_type_returns_403(self):
        """Requesting an unknown export type should return HTTP 403."""
        response = self._get_export("unknown_type_xyz")
        self.assertEqual(response.status_code, 403)

    def test_export_requires_login(self):
        """Unauthenticated user should not access export endpoint."""
        self.client.logout()
        response = self._get_export("orders")
        self.assertIn(response.status_code, [302, 403])


# ---------------------------------------------------------------------------
# P&L / GROSS MARGIN REPORT — refund-netting regression test
#
# gross_margin_report used to sum Order.grand_total, which never changes after
# a refund (approve_refund records the refund as a separate negative Payment
# instead). That overstated "Gross Revenue" by the refunded amount and
# disagreed with the Sales Dashboard (daily_sales), which already nets
# refunds. Fixed to sum actual Payment rows for the order instead.
# ---------------------------------------------------------------------------

class GrossMarginReportRefundTest(TestCase):
    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()

    def test_gross_revenue_nets_a_refund(self):
        from reports.services.pl_reports import gross_margin_report

        order = create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=500)
        # A partial refund against that same order — recorded as its own
        # negative Payment row, exactly like approve_refund does in production.
        Payment.objects.create(
            order=order, method="refund", amount=Decimal("-150.00"),
            reference="REFUND-1", created_by=self.user,
        )

        today = timezone.localdate()
        result = gross_margin_report(self.tenant, self.outlet, today, today)

        # 500 collected - 150 refunded = 350 net, NOT 500.
        self.assertEqual(result["gross_revenue"], 350.0)

    def test_gross_revenue_matches_full_payment_with_no_refund(self):
        from reports.services.pl_reports import gross_margin_report

        create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=500)

        today = timezone.localdate()
        result = gross_margin_report(self.tenant, self.outlet, today, today)

        self.assertEqual(result["gross_revenue"], 500.0)


# ---------------------------------------------------------------------------
# owner_dashboard_metrics business-date tests
#
# Same bug class as the Z-report fix — a single created_at__date= filter
# against timezone.localdate() ignores the outlet's business-day cutoff,
# so any order placed after midnight but before the cutoff hour would
# vanish from the live dashboard until the calendar caught up.
# ---------------------------------------------------------------------------

from datetime import datetime as _dt
from unittest.mock import patch


class OwnerDashboardMetricsBusinessDateTest(TestCase):

    def setUp(self):
        self.tenant = Tenant.objects.create(name="Dashboard BD Tenant")
        self.outlet = Outlet.objects.create(
            tenant=self.tenant, name="Main", business_day_start_hour=6
        )
        self.owner = User.objects.create_user(
            username="dash_owner", password="pw", tenant=self.tenant,
            outlet=self.outlet, role="owner",
        )

    def _order_at(self, naive_dt, amount):
        order = Order.objects.create(
            tenant=self.tenant, outlet=self.outlet, status="paid",
            subtotal=Decimal(amount), grand_total=Decimal(amount),
        )
        aware = timezone.make_aware(naive_dt, timezone.get_current_timezone())
        Order.objects.filter(id=order.id).update(created_at=aware)
        Payment.objects.create(order=order, method="cash", amount=Decimal(amount))
        return order

    @patch("django.utils.timezone.now")
    def test_late_night_order_counted_on_dashboard(self, mock_now):
        from reports.services.dashboard_metrics import owner_dashboard_metrics

        mock_now.return_value = timezone.make_aware(
            _dt(2026, 7, 17, 21, 0), timezone.get_current_timezone()
        )
        evening_order = self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        # Past midnight, before the 6 AM cutoff — same business day.
        late_order = self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        # Owner checks the dashboard at 4 AM, still the same business day
        # as both orders above.
        mock_now.return_value = timezone.make_aware(
            _dt(2026, 7, 18, 4, 0), timezone.get_current_timezone()
        )

        results = owner_dashboard_metrics(self.owner)
        self.assertEqual(len(results), 1)
        # Under the old bug, a plain created_at__date filter evaluated at
        # 4 AM would only find the 2 AM order (₹250), dropping the whole
        # prior evening's ₹1000.
        self.assertEqual(float(results[0]["revenue"]), 1250.0)
        self.assertEqual(results[0]["orders"], 2)


# ---------------------------------------------------------------------------
# export_services.py business-date tests
#
# A separate module from the report *services* fixed earlier — this file
# has its own independent date-filtering, missed in the first sweep, found
# by specifically re-checking GSTR-1 for accuracy. Same bug, same fix,
# applied here across every generate_*_csv/excel function, GSTR-1 included
# since it's the one that actually gets filed with the government.
# ---------------------------------------------------------------------------

class ExportServicesBusinessDateTest(TestCase):

    def setUp(self):
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()
        self.outlet.business_day_start_hour = 6
        self.outlet.save(update_fields=["business_day_start_hour"])

    def _order_at(self, naive_dt, amount):
        order = create_paid_order(self.tenant, self.outlet, self.user, self.item, grand_total=amount)
        aware = timezone.make_aware(naive_dt, timezone.get_current_timezone())
        Order.objects.filter(id=order.id).update(created_at=aware)
        return order

    def test_gstr1_includes_late_night_order_in_correct_business_day(self):
        from datetime import date
        from reports.services.export_services import generate_gstr1_excel
        from openpyxl import load_workbook

        # 9 PM order (business day = the 17th) and a 2 AM order the *next*
        # calendar day that's still the same business day under a 6 AM
        # cutoff — the exact case a plain created_at__date filter drops.
        self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        business_date = date(2026, 7, 17)
        result = generate_gstr1_excel(self.tenant, self.outlet, business_date, business_date)

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        total_row = [row for row in ws.iter_rows(values_only=True) if row and row[0] == "TOTAL"][0]
        taxable_total = total_row[3]

        # create_paid_order's fixture item always carries a fixed
        # total_price of 400 regardless of the order's grand_total — so
        # both orders' item lines (400 + 400 = 800) must be present. Under
        # the old bug, only the 2 AM order's line (400) would show up.
        self.assertEqual(taxable_total, 800.0)

    def test_orders_csv_includes_late_night_order_in_correct_business_day(self):
        from datetime import date
        from reports.services.export_services import generate_orders_csv

        self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        business_date = date(2026, 7, 17)
        result = generate_orders_csv(self.tenant, self.outlet, business_date, business_date)
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)[1:]  # skip header

        self.assertEqual(len(rows), 2)

    def test_items_csv_includes_late_night_order_in_correct_business_day(self):
        from datetime import date
        from reports.services.export_services import generate_items_csv

        self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        business_date = date(2026, 7, 17)
        result = generate_items_csv(self.tenant, self.outlet, business_date, business_date)
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)[1:]  # skip header

        # Both orders share the same fixture item, so this collapses into
        # one row — quantity should reflect both orders (2 each = 4), not
        # just the 2 AM one.
        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0][2]), 4)  # Quantity Sold column

    def test_waiter_csv_includes_late_night_order_in_correct_business_day(self):
        from datetime import date
        from reports.services.export_services import generate_waiter_csv

        self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        business_date = date(2026, 7, 17)
        result = generate_waiter_csv(self.tenant, self.outlet, business_date, business_date)
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)[1:]  # skip header

        # Both orders were created_by the same fixture user, so this
        # collapses into one row — 2 orders, not just the 2 AM one.
        self.assertEqual(len(rows), 1)
        self.assertEqual(int(rows[0][1]), 2)  # Total Orders Handled column

    def test_category_csv_includes_late_night_order_in_correct_business_day(self):
        from datetime import date
        from reports.services.export_services import generate_category_csv

        self._order_at(_dt(2026, 7, 17, 21, 0), "1000.00")
        self._order_at(_dt(2026, 7, 18, 2, 0), "250.00")

        business_date = date(2026, 7, 17)
        result = generate_category_csv(self.tenant, self.outlet, business_date, business_date)
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)[1:]  # skip header

        # Same fixture item/category on both orders — one row, revenue
        # from both orders' item lines (400 + 400), not just the 2 AM one.
        self.assertEqual(len(rows), 1)
        self.assertEqual(float(rows[0][2]), 800.0)  # Total Revenue column


# ---------------------------------------------------------------------------
# "POS Dashboard" nav link on the Kitchen KPIs page used to hardcode
# /tables/ regardless of whether the tenant has a floor plan at all. A
# default "cafe" tenant_type gets qr_menu + kitchen_display but NOT
# floor_plan (core/features.py) -- landing on /tables/ there 403s or shows
# an unusable empty floor plan. Should go to /token/ instead.
# ---------------------------------------------------------------------------

class KitchenDashboardPosLinkTest(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Tableless Cafe", tenant_type="cafe")
        self.outlet = Outlet.objects.create(tenant=self.tenant, name="Main")
        self.owner = User.objects.create_user(
            username="cafe_owner", password="pw", tenant=self.tenant,
            outlet=self.outlet, role="owner",
        )
        self.client = Client()
        self.client.login(username="cafe_owner", password="pw")

    def test_pos_dashboard_link_points_to_token_not_tables(self):
        response = self.client.get("/reports/kitchen/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('href="/token/"', content)
        self.assertNotIn('href="/tables/"', content)

    def test_header_qr_approval_icon_absent_for_tenant_without_floor_plan(self):
        # Same base.html header renders on this page too. The icon is gated
        # on qr_menu AND floor_plan now -- a cafe without floor_plan has no
        # waiter-run floor to send this to, and Token Billing already shows
        # review items inline with its own approve/reject buttons, so the
        # icon (not just its link) is hidden entirely for this tenant.
        response = self.client.get("/reports/kitchen/")
        content = response.content.decode()
        self.assertNotIn('title="QR Orders awaiting approval"', content)

    def test_header_qr_approval_icon_present_for_tenant_with_floor_plan(self):
        floor_tenant = Tenant.objects.create(name="Floor Plan Cafe", tenant_type="cafe")
        floor_outlet = Outlet.objects.create(tenant=floor_tenant, name="Main")
        from tenants.models import TenantFeatureOverride
        TenantFeatureOverride.objects.create(tenant=floor_tenant, feature="floor_plan", enabled=True)
        owner = User.objects.create_user(
            username="floor_owner", password="pw", tenant=floor_tenant,
            outlet=floor_outlet, role="owner",
        )
        client = Client()
        client.login(username="floor_owner", password="pw")
        response = client.get("/reports/kitchen/")
        content = response.content.decode()
        self.assertIn('<a href="/tables/" class="btn-icon" title="QR Orders awaiting approval">', content)


class ReportTabsConsistencyTest(TestCase):
    """
    All seven report pages used to hand-copy the same tab strip (six of
    them), or reimplement it a third way with a different CSS class
    entirely (kitchen_dashboard.html), or skip it altogether
    (inventory_report.html was a fully standalone document with no shared
    header at all). Drift had already happened silently: audit_report.html
    was missing the CRM Analytics tab entirely, and inventory_report.html's
    own copy of the dark-mode toggle stored the preference under a
    different value format than every other page, silently corrupting the
    shared preference. Now every page renders the same
    reports/_report_tabs.html include, so this class exists to make sure
    that stays true.
    """

    REPORT_URLS = [
        "/reports/dashboard/",
        "/reports/kitchen/",
        "/reports/inventory/",
        "/reports/menu-engineering/",
        "/reports/labor/",
        "/reports/audit/",
        "/reports/crm-analytics/",
    ]

    ALL_TAB_LABELS = [
        "Sales", "Kitchen KPIs", "Inventory", "Menu Engineering",
        "Labor Cost", "Discount/Void Audit", "CRM Analytics",
    ]

    def setUp(self):
        self.client = Client()
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()
        from tenants.models import TenantFeatureOverride
        TenantFeatureOverride.objects.create(tenant=self.tenant, feature="advanced_reports", enabled=True)
        TenantFeatureOverride.objects.create(tenant=self.tenant, feature="crm", enabled=True)
        self.client.login(username="owner_test", password="pass1234")

    def test_every_report_page_shows_all_seven_tabs(self):
        for url in self.REPORT_URLS:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
                content = response.content.decode()
                for label in self.ALL_TAB_LABELS:
                    self.assertIn(label, content, f"{url} is missing the '{label}' tab")

    def test_inventory_tab_marked_active_on_inventory_report(self):
        response = self.client.get("/reports/inventory/")
        content = response.content.decode()
        self.assertIn('href="/reports/inventory/" class="report-tab active">Inventory', content)

    def test_kitchen_tab_marked_active_on_kitchen_dashboard(self):
        response = self.client.get("/reports/kitchen/")
        content = response.content.decode()
        self.assertIn('href="/reports/kitchen/" class="report-tab active">Kitchen KPIs', content)

    def test_audit_tab_marked_active_on_audit_report(self):
        response = self.client.get("/reports/audit/")
        content = response.content.decode()
        self.assertIn('href="/reports/audit/" class="report-tab active">Discount/Void Audit', content)


class InventoryReportSharedHeaderTest(TestCase):
    """
    inventory_report.html used to be a fully standalone HTML document: its
    own <html>/<head>, its own hardcoded color/font variables ignoring
    whatever the rest of the app resolves for a tenant, and its own
    dark-mode toggle that stored localStorage['dark'] as '1'/'0' instead
    of the 'true'/'false' every other page (core/base.html) uses --
    meaning toggling dark mode on this one page silently broke it
    everywhere else in the app. It now extends core/base.html like every
    other report page.
    """

    def setUp(self):
        self.client = Client()
        self.tenant, self.outlet, self.user, self.item = create_base_fixtures()
        self.client.login(username="owner_test", password="pass1234")

    def test_uses_the_shared_app_header(self):
        response = self.client.get("/reports/inventory/")
        content = response.content.decode()
        self.assertIn('class="pos-header"', content)
        self.assertIn("themeManager.toggle()", content)

    def test_no_longer_has_its_own_incompatible_dark_mode_storage(self):
        response = self.client.get("/reports/inventory/")
        content = response.content.decode()
        self.assertNotIn("localStorage.setItem('dark'", content)
        self.assertNotIn('localStorage.getItem(\'dark\')', content)
        self.assertNotIn("body.dark {", content)

    def test_report_content_still_renders(self):
        """The actual point of the page -- make sure the shell swap didn't
        drop any of the real report content."""
        response = self.client.get("/reports/inventory/")
        content = response.content.decode()
        self.assertIn("Stock Ledger", content)
        self.assertIn("Consumption", content)
        self.assertIn("Wastage", content)
        self.assertIn("Cost Analysis", content)
        self.assertIn("Can Make", content)
        self.assertIn("Closing Stock", content)